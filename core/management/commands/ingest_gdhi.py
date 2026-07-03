"""
Ingest ONS regional GDHI (gross disposable household income) at LAD.

Dataset: "Regional gross disposable household income: local authorities by ITL1
region" — the same ONS workbook template as the GVA files (one workbook per ITL1
region). Each workbook's summary sheets are clean wide tables (Region, LAD code,
Region name, year columns):

    Table 1  — total GDHI, £ million   -> gdhi-total   (CURRENCY, additive)
    Table 3  — GDHI per head, £         -> gdhi-per-head (RATIO, non-additive)

Sheets are selected by their title (not a hard-coded number), so table renumbering
between editions won't silently pick the wrong one. Unlike GVA, per head is NOT
derived — ONS publishes it directly (residence-based, mid-year population), so we
ingest ONS's figure as its own single-source observation.

    python manage.py ingest_gdhi --path seed_data/gdhi/
    python manage.py ingest_gdhi --path <dir-or-file> --dry-run
    python manage.py ingest_gdhi --url <one-region.xlsx> --vintage 2024-09-04

period_type = CALENDAR_YEAR; the latest year in each file is marked PROVISIONAL.
Vintage is auto-detected from the Information sheet ("published on <date>").
"""

import glob
import io
import os
import re
import urllib.request
from datetime import date, datetime
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from core.models import (
    Indicator,
    IndicatorDomain,
    ObservationStatus,
    PeriodType,
    Place,
    PlaceObservation,
    PlaceTier,
    Source,
    SubjectScope,
    ValueType,
)

FILE_GLOB = "*.xlsx"
SOURCE_NAME = "ONS Regional accounts (GVA / GDP / GDHI)"
SOURCE_PUBLISHER = "Office for National Statistics"
TOTAL_CODE = "gdhi-total"
PERHEAD_CODE = "gdhi-per-head"


def _title(ws):
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return str(row[0]).lower() if row and row[0] else ""
    return ""


def _pick_sheet(wb, want):
    """Find the total-£m or per-head-£ summary sheet by its title."""
    for sn in wb.sheetnames:
        t = _title(wb[sn])
        if want == "total" and "total gdhi" in t and "million" in t:
            return sn
        if (want == "perhead" and "per head" in t and "pound" in t
                and "indic" not in t and "index" not in t):
            return sn
    return None


def _detect_vintage(wb):
    if "Information" not in wb.sheetnames:
        return None
    for row in wb["Information"].iter_rows(max_row=30, values_only=True):
        for cell in row:
            if not cell:
                continue
            m = re.search(r"published on (\d{1,2} \w+ \d{4})", str(cell))
            if m:
                try:
                    return datetime.strptime(m.group(1), "%d %B %Y").date().isoformat()
                except ValueError:
                    return m.group(1)
    return None


def parse_table(ws):
    """Yield (gss_code, name, year, Decimal value, is_latest_year)."""
    allrows = list(ws.iter_rows(values_only=True))
    header, hi = None, None
    for i, r in enumerate(allrows[:10]):
        if any(str(c).strip() == "LAD code" for c in r if c is not None):
            header = [str(c).strip() if c is not None else "" for c in r]
            hi = i
            break
    if header is None:
        raise CommandError(f"{ws.title}: no header row with 'LAD code'.")

    code_col = header.index("LAD code")
    name_col = header.index("Region name") if "Region name" in header else code_col + 1
    years = {j: int(h) for j, h in enumerate(header) if h.isdigit() and 1990 <= int(h) <= 2100}
    if not years:
        raise CommandError(f"{ws.title}: no year columns found.")
    max_year = max(years.values())

    for r in allrows[hi + 1:]:
        code = r[code_col] if code_col < len(r) else None
        if not code:
            continue
        name = r[name_col] if name_col < len(r) else ""
        for j, yr in years.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)):
                yield (str(code).strip(), str(name).strip() if name else "",
                       yr, Decimal(str(v)), yr == max_year)


class Command(BaseCommand):
    help = "Ingest ONS regional GDHI (total £m + per head £) at LAD into PlaceObservation."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None, help="Directory of region files, or a single file.")
        parser.add_argument("--url", default=None, help="URL of a single region .xlsx.")
        parser.add_argument("--vintage", default=None, help="Override the auto-detected release vintage.")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        sources = self._workbook_sources(opts)
        if not sources:
            raise CommandError("Provide --path (dir or file) or --url.")

        total_ind, perhead_ind = self._indicators()
        source, _ = Source.objects.get_or_create(name=SOURCE_NAME, publisher=SOURCE_PUBLISHER)
        place_by_code = {
            p.gss_code: p
            for p in Place.objects.filter(tier=PlaceTier.LAD).order_by("valid_from")
        }

        objs, unmatched = [], {}
        counts = {TOTAL_CODE: 0, PERHEAD_CODE: 0}
        vintages = set()

        for label, wb in sources:
            vintage = opts["vintage"] or _detect_vintage(wb) or "unknown"
            vintages.add(vintage)
            for want, ind, code in (("total", total_ind, TOTAL_CODE),
                                    ("perhead", perhead_ind, PERHEAD_CODE)):
                sn = _pick_sheet(wb, want)
                if sn is None:
                    self.stdout.write(self.style.WARNING(f"{label}: no {want} sheet found — skipped."))
                    continue
                for gss, name, yr, val, is_latest in parse_table(wb[sn]):
                    place = place_by_code.get(gss)
                    if place is None:
                        unmatched[gss] = name
                        continue
                    if not opts["dry_run"]:
                        objs.append(PlaceObservation(
                            indicator=ind, place=place,
                            period_start=date(yr, 1, 1), period_end=date(yr, 12, 31),
                            period_type=PeriodType.CALENDAR_YEAR,
                            value=val, unit="", source=source, vintage=vintage,
                            status=ObservationStatus.PROVISIONAL if is_latest else "",
                        ))
                    counts[code] += 1

        if opts["dry_run"]:
            self.stdout.write(
                f"GDHI dry run: {counts[TOTAL_CODE]} total rows, {counts[PERHEAD_CODE]} per-head rows, "
                f"vintage(s)={sorted(vintages)}, {len(unmatched)} unmatched codes."
            )
            return

        before = PlaceObservation.objects.filter(
            indicator__in=[total_ind, perhead_ind], source=source,
        ).count()
        with transaction.atomic():
            for i in range(0, len(objs), 5000):
                PlaceObservation.objects.bulk_create(objs[i:i + 5000], ignore_conflicts=True)
        after = PlaceObservation.objects.filter(
            indicator__in=[total_ind, perhead_ind], source=source,
        ).count()

        self.stdout.write(self.style.SUCCESS(
            f"GDHI: {after - before} created "
            f"(gdhi-total {counts[TOTAL_CODE]} rows, gdhi-per-head {counts[PERHEAD_CODE]} rows), "
            f"vintage(s) {sorted(vintages)}."
        ))
        if unmatched:
            self.stdout.write(self.style.WARNING(
                f"{len(unmatched)} codes had no matching LAD Place (region rows / churn). Examples:"
            ))
            for code, name in list(unmatched.items())[:8]:
                self.stdout.write(f"  {code}  {name}")

    # -- helpers -----------------------------------------------------------

    def _workbook_sources(self, opts):
        """Return [(label, workbook), ...] from --path (dir/file) or --url."""
        paths = []
        if opts["path"]:
            if os.path.isdir(opts["path"]):
                paths = sorted(glob.glob(os.path.join(opts["path"], FILE_GLOB)))
            elif os.path.isfile(opts["path"]):
                paths = [opts["path"]]
            else:
                raise CommandError(f"--path not found: {opts['path']!r}")
        result = []
        for p in paths:
            result.append((os.path.basename(p), load_workbook(p, read_only=True, data_only=True)))
        if opts["url"]:
            self.stdout.write(f"Fetching {opts['url']} …")
            data = urllib.request.urlopen(opts["url"], timeout=180).read()
            result.append((opts["url"], load_workbook(io.BytesIO(data), read_only=True, data_only=True)))
        return result

    @staticmethod
    def _indicators():
        economy, _ = IndicatorDomain.objects.get_or_create(code="economy", defaults={"name": "Economy"})
        total, _ = Indicator.objects.get_or_create(
            code=TOTAL_CODE,
            defaults=dict(name="Gross disposable household income, total", domain=economy,
                          unit="£m", value_type=ValueType.CURRENCY, is_additive=True,
                          subject_scope=SubjectScope.PLACE),
        )
        perhead, _ = Indicator.objects.get_or_create(
            code=PERHEAD_CODE,
            defaults=dict(name="Gross disposable household income per head", domain=economy,
                          unit="£", value_type=ValueType.RATIO, is_additive=False,
                          subject_scope=SubjectScope.PLACE),
        )
        return total, perhead
