"""
Ingest Scottish deprivation (SIMD) at the LAD tier from the Data-Zone-level ranks file.

Scotland only: the Scottish Index of Multiple Deprivation is its own statistical system —
deprivation is modelled per nation and NEVER merged into a UK-wide indicator, and never
compared across nations. SIMD is RANK-based with NO published composite score, so only a
decile-share metric is derived (a synthesised score from domain ranks would be editorial
invention — explicitly out of scope).

    python manage.py ingest_simd                    # fetch SIMD 2020v2 ranks from gov.scot
    python manage.py ingest_simd --path ranks.xlsx  # or a local copy
    python manage.py ingest_simd --dry-run

One LAD metric is derived from the Data Zone rows (Data Zones are NOT Place rows — we only
aggregate through them; raw Data Zone ranks are never stored):

  simd-most-deprived-decile-share-scotland
      % of a council area's Data Zones that fall in Scotland's most-deprived decile
      (within-Scotland decile). RATE, non-additive.

The decile is DERIVED from the overall rank — the file carries no decile column. gov.scot
defines decile 1 as the most-deprived 10% of Data Zones; for N Data Zones that is the
ceil(N/10) lowest-ranked (the ntile "larger groups first" convention gov.scot uses). This
was validated against gov.scot's own published local-authority figures before shipping:
Inverclyde's 20%-most-deprived share reproduces exactly (51/114 = 44.7%) and Glasgow's
deciles 1-3 count reproduces exactly (422 zones) — see CLAUDE.md.

Council area is a NAME in the file, not a GSS code, so it is matched to the spine's 32 S12
Places by name, with one alias for the Western Isles ("Na h-Eileanan an Iar" in the file
vs "Na h-Eileanan Siar" in the spine). All 32 councils resolve.

period: POINT at the edition date. source: Scottish Index of Multiple Deprivation (seeded).
vintage: the edition (SIMD2020v2); a later edition loads as a new vintage + period.
"""

import io
import math
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from core.models import (
    Indicator,
    ObservationStatus,
    PeriodType,
    Place,
    PlaceObservation,
    PlaceTier,
    Source,
)

SIMD_RANKS_URL = (
    "https://www.gov.scot/binaries/content/documents/govscot/publications/statistics/"
    "2020/01/scottish-index-of-multiple-deprivation-2020-ranks-and-domain-ranks/documents/"
    "scottish-index-of-multiple-deprivation-2020-ranks-and-domain-ranks/"
    "scottish-index-of-multiple-deprivation-2020-ranks-and-domain-ranks/govscot%3Adocument/"
    "SIMD%2B2020v2%2B-%2Branks.xlsx"
)
SHEET = "SIMD 2020v2 ranks"
COL_DATAZONE = "Data_Zone"
COL_COUNCIL = "Council_area"
COL_RANK = "SIMD2020v2_Rank"
INDICATOR_CODE = "simd-most-deprived-decile-share-scotland"
SOURCE_NAME = "Scottish Index of Multiple Deprivation"
SOURCE_PUBLISHER = "Scottish Government"

# The file names the Western Isles "Na h-Eileanan an Iar"; the spine uses the ONS/GSS
# name "Na h-Eileanan Siar" (S12000013). Same council — an explicit alias, not a
# normalisation (the two Gaelic renderings differ genuinely). All 32 councils resolve.
COUNCIL_ALIASES = {"Na h-Eileanan an Iar": "Na h-Eileanan Siar"}


def parse_simd_ranks(path_or_fileobj):
    """Yield (data_zone, council_name, rank) from the SIMD ranks sheet."""
    wb = load_workbook(path_or_fileobj, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise CommandError(f"SIMD workbook has no sheet {SHEET!r} (has {wb.sheetnames}).")
    ws = wb[SHEET]
    header = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header = [str(c).strip() if c is not None else "" for c in row]
        break
    for col in (COL_DATAZONE, COL_COUNCIL, COL_RANK):
        if col not in header:
            raise CommandError(f"SIMD sheet missing column {col!r} (edition changed?).")
    dz_j, c_j, r_j = header.index(COL_DATAZONE), header.index(COL_COUNCIL), header.index(COL_RANK)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dz = row[dz_j] if dz_j < len(row) else None
        if not dz:
            continue
        yield str(dz).strip(), str(row[c_j]).strip(), int(row[r_j])


class Command(BaseCommand):
    help = "Ingest SIMD most-deprived-decile-share at the LAD tier (Data Zone -> council)."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None, help="Local ranks .xlsx (else fetch from gov.scot).")
        parser.add_argument("--edition-date", default="2020-01-28", help="POINT period for this edition.")
        parser.add_argument("--vintage", default="SIMD2020v2")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        edition = datetime.strptime(opts["edition_date"], "%Y-%m-%d").date()
        vintage = opts["vintage"]

        try:
            indicator = Indicator.objects.get(code=INDICATOR_CODE)
        except Indicator.DoesNotExist:
            raise CommandError(f"Indicator {INDICATOR_CODE!r} not seeded — run migrations / seed_v1.")
        source, _ = Source.objects.get_or_create(name=SOURCE_NAME, publisher=SOURCE_PUBLISHER)

        src = opts["path"] or self._fetch()
        rows = list(parse_simd_ranks(src))
        if not rows:
            raise CommandError("SIMD ranks file parsed to zero rows.")

        # National most-deprived decile = the ceil(N/10) lowest-ranked Data Zones
        # (gov.scot "most deprived 10%", ntile larger-groups-first). Derived from rank —
        # the file has no decile column.
        n = len(rows)
        decile1_cut = math.ceil(n / 10)

        # Aggregate Data Zones -> council (raw ranks never stored, only the counts).
        agg = defaultdict(lambda: [0, 0])   # council_name -> [n_dz, n_in_decile1]
        for _dz, council, rank in rows:
            a = agg[council]
            a[0] += 1
            if rank <= decile1_cut:
                a[1] += 1

        # Council NAME -> S12 Place (spine names, plus the one alias).
        place_by_name = {
            p.name: p
            for p in Place.objects.filter(tier=PlaceTier.LAD, gss_code__startswith="S12")
        }
        objs, unmatched = [], []
        for council, (total, d1) in agg.items():
            place = place_by_name.get(COUNCIL_ALIASES.get(council, council))
            if place is None:
                unmatched.append(council)
                continue
            share = (Decimal(d1) / total * 100).quantize(Decimal("0.01"))
            objs.append(PlaceObservation(
                indicator=indicator, place=place,
                period_start=edition, period_end=edition, period_type=PeriodType.POINT,
                value=share, unit="", source=source, vintage=vintage,
                status=ObservationStatus.FINAL,
            ))

        if opts["dry_run"]:
            self.stdout.write(
                f"SIMD dry run: {n} Data Zones, decile-1 cut rank<={decile1_cut}, "
                f"{len(agg)} councils, {len(objs)} matched, {len(unmatched)} unmatched "
                f"(vintage {vintage}, {edition}).")
            if unmatched:
                self.stdout.write(f"  unmatched: {unmatched}")
            return

        if unmatched:
            # All 32 councils are expected to resolve — an unmatched one means the file's
            # council naming drifted. Fail loudly rather than silently drop a whole council.
            raise CommandError(f"{len(unmatched)} council(s) did not match an S12 Place: "
                               f"{unmatched}. Add an alias in COUNCIL_ALIASES.")

        before = PlaceObservation.objects.filter(
            indicator=indicator, source=source, vintage=vintage).count()
        with transaction.atomic():
            PlaceObservation.objects.bulk_create(objs, ignore_conflicts=True)
        after = PlaceObservation.objects.filter(
            indicator=indicator, source=source, vintage=vintage).count()

        self.stdout.write(self.style.SUCCESS(
            f"SIMD (Scotland): {after - before} created ({len(objs)} councils, "
            f"decile-1 cut rank<={decile1_cut} of {n} Data Zones), vintage {vintage}, {edition}."))

    def _fetch(self):
        self.stdout.write(f"Fetching SIMD ranks from {SIMD_RANKS_URL} ...")
        try:
            with urllib.request.urlopen(
                    urllib.request.Request(SIMD_RANKS_URL, headers={"User-Agent": "Mozilla/5.0"}),
                    timeout=180) as resp:
                data = resp.read()
        except Exception as exc:
            raise CommandError(f"Could not fetch SIMD ranks file: {exc}")
        if len(data) < 10000:
            raise CommandError(f"SIMD download too small ({len(data)} bytes) — URL moved?")
        return io.BytesIO(data)
