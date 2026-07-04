"""
Ingest life expectancy at birth for ALL FOUR UK nations from the single ONS
"Life expectancy for local areas of the UK" dataset.

Place at: core/management/commands/ingest_le_ons.py

Why this exists alongside ingest_fingertips: the Fingertips loader covers England
only. This ONS release publishes England, Wales, Scotland AND Northern Ireland in
one file on ONE uniform methodology, which is exactly what makes cross-nation LE
comparison honest. So we re-source all four nations from here (Option B): it lands
as a NEW vintage beside the Fingertips England vintage (append-only — nothing is
overwritten), and latest-vintage-per-period then shows the ONS series everywhere.

Verified before building (see CLAUDE.md): England ONS vs the current Fingertips
values are IDENTICAL to 0.0 years across sampled LADs at the latest common period
(Fingertips 90366 is sourced from ONS's own calculation), so there is no
methodology gap — no reason to keep England on a separate source. W/S/N area codes
match the spine exactly (Wales 22/22, Scotland 32/32, NI 11/11).

    python manage.py ingest_le_ons                 # fetch live from ONS
    python manage.py ingest_le_ons --dry-run
    python manage.py ingest_le_ons --path le.xlsx  # use a local copy

Key shape decisions:
- Extends the EXISTING life-expectancy-birth-male / -female indicators (LE is
  comparable across nations, unlike deprivation — no per-nation indicators).
- Sheet "1" is the full long-format table (header on row 6). We keep Area type
  "Local Areas" only (drops Country / Region / Combined Authority / Health Board /
  ICB comparator rows), Age group "<1" (life expectancy AT BIRTH), and Sex
  Male / Female (the series is published by sex only — no "Persons").
- Period "2001 to 2003" -> period_start 2001-01-01, period_end 2003-12-31, the
  3-year pooled series (matches the Fingertips period convention exactly).
- Area codes match the LAD spine across all four nations (E06/E07/E08/E09 +
  W06 / S12 / N09). English counties (E10, upper-tier) sit in "Local Areas" too
  but have no LAD Place — reported as unmatched like any geography drift, never
  force-fitted (they would double-count against their districts in a roll-up).
"""

import io
import urllib.request
from datetime import date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from core.models import (
    Indicator,
    ObservationStatus,
    PeriodType,
    Place,
    PlaceObservation,
    PlaceTier,
    Source,
)

# Edition-specific URL (the ONS file path encodes the year range). If a later
# edition moves it, pass --path with a downloaded copy or update this constant.
ONS_URL = (
    "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/"
    "healthandsocialcare/healthandlifeexpectancies/datasets/"
    "lifeexpectancyforlocalareasoftheuk/"
    "between2001to2003and2022to2024/lifeexpectancylocalareas.xlsx"
)
DATA_SHEET = "1"                 # full long-format table; header on row 6
AREA_TYPE_KEEP = "Local Areas"   # drops Country/Region/Combined Authority/Health Board/ICB
AGE_AT_BIRTH = "<1"              # life expectancy AT BIRTH
DEFAULT_VINTAGE = "ons-le-2025-12-10"   # release date on the Cover sheet
SOURCE_NAME = "ONS life expectancy for local areas"
SOURCE_PUBLISHER = "Office for National Statistics"

# Pure recodes: this ONS product emits new GSS codes for two metropolitan districts that
# standard ONS LAD geography (and every other source) still codes the old way, so the spine
# keeps the old code. Map the new code back to the existing Place so the LE lands on it.
# Same real place, new identifier — an alias, not a restructure (see refresh_lad_spine).
RECODE_ALIASES = {
    "E08000038": "E08000016",   # Barnsley
    "E08000039": "E08000019",   # Sheffield
}

# ONS "Sex" value -> our existing indicator code (published by sex only, no Persons).
SEX_TO_CODE = {
    "Male": "life-expectancy-birth-male",
    "Female": "life-expectancy-birth-female",
}


def _pooled_period(text):
    """'2001 to 2003' -> (date(2001,1,1), date(2003,12,31)); None if not a range."""
    parts = str(text).strip().split(" to ")
    if len(parts) != 2 or not (parts[0].isdigit() and parts[1].isdigit()):
        return None
    return date(int(parts[0]), 1, 1), date(int(parts[1]), 12, 31)


def parse_le(path_or_fileobj):
    """Yield (area_code, area_name, sex, period_start, period_end, Decimal value)
    for Local-Areas, at-birth, Male/Female rows of the ONS LE workbook."""
    wb = load_workbook(path_or_fileobj, read_only=True, data_only=True)
    if DATA_SHEET not in wb.sheetnames:
        raise CommandError(f"ONS LE workbook has no sheet {DATA_SHEET!r} "
                           f"(has {wb.sheetnames}).")
    ws = wb[DATA_SHEET]

    header, hrow = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Area code" in cells and "Life expectancy" in cells and "Sex" in cells:
            header, hrow = cells, i
            break
    if header is None:
        raise CommandError("ONS LE workbook: could not locate the header row on "
                           f"sheet {DATA_SHEET!r}.")
    col = {name: j for j, name in enumerate(header)}

    def cell(row, name):
        j = col.get(name)
        return row[j] if j is not None and j < len(row) else None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hrow:
            continue
        if str(cell(row, "Area type") or "").strip() != AREA_TYPE_KEEP:
            continue
        if str(cell(row, "Age group") or "").strip() != AGE_AT_BIRTH:
            continue
        sex = str(cell(row, "Sex") or "").strip()
        if sex not in SEX_TO_CODE:
            continue
        period = _pooled_period(cell(row, "Period"))
        if period is None:
            continue
        try:
            value = Decimal(str(cell(row, "Life expectancy")).strip())
        except (InvalidOperation, ValueError, AttributeError):
            continue
        yield (str(cell(row, "Area code")).strip(),
               str(cell(row, "Area name") or "").strip(),
               sex, period[0], period[1], value)


class Command(BaseCommand):
    help = "Ingest ONS life expectancy at birth (all four UK nations, LAD tier)."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None,
                            help="Local .xlsx path; defaults to fetching live from ONS.")
        parser.add_argument("--vintage", default=DEFAULT_VINTAGE,
                            help=f"Release vintage (default {DEFAULT_VINTAGE}).")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        vintage = opts["vintage"]

        indicators = {}
        for code in SEX_TO_CODE.values():
            try:
                indicators[code] = Indicator.objects.get(code=code)
            except Indicator.DoesNotExist:
                raise CommandError(f"Indicator {code!r} not seeded — run migrations / seed_v1.")
        source, _ = Source.objects.get_or_create(name=SOURCE_NAME, publisher=SOURCE_PUBLISHER)

        src = opts["path"] or self._fetch()
        rows = list(parse_le(src))

        # LAD places by code across ALL nations (later valid_from wins for the
        # current boundary set — LADs are single-version here).
        place_by_code = {
            p.gss_code: p for p in Place.objects.filter(tier=PlaceTier.LAD).order_by("valid_from")
        }

        objs, unmatched, by_nation, kept = [], {}, {}, 0
        for code, name, sex, ps, pe, value in rows:
            code = RECODE_ALIASES.get(code, code)   # Barnsley/Sheffield recodes -> spine code
            place = place_by_code.get(code)
            if place is None:
                unmatched[code] = name
                continue
            kept += 1
            by_nation[place.nation] = by_nation.get(place.nation, 0) + 1
            objs.append(PlaceObservation(
                indicator=indicators[SEX_TO_CODE[sex]], place=place,
                period_start=ps, period_end=pe, period_type=PeriodType.CALENDAR_YEAR,
                value=value, unit="", source=source, vintage=vintage,
                status=ObservationStatus.FINAL,
            ))

        if opts["dry_run"]:
            self.stdout.write(
                f"ONS LE dry run: {kept} observations "
                f"({', '.join(f'{n}={c}' for n, c in sorted(by_nation.items()))}), "
                f"{len(unmatched)} unmatched areas, vintage {vintage}.")
            return

        before = PlaceObservation.objects.filter(
            indicator__in=indicators.values(), source=source, vintage=vintage).count()
        with transaction.atomic():
            for i in range(0, len(objs), 5000):
                PlaceObservation.objects.bulk_create(objs[i:i + 5000], ignore_conflicts=True)
        after = PlaceObservation.objects.filter(
            indicator__in=indicators.values(), source=source, vintage=vintage).count()

        self.stdout.write(self.style.SUCCESS(
            f"ONS life expectancy: {after - before} created "
            f"({', '.join(f'{n}={c}' for n, c in sorted(by_nation.items()))}), "
            f"vintage {vintage}."))
        if unmatched:
            self.stdout.write(self.style.WARNING(
                f"{len(unmatched)} areas had no matching LAD Place (upper-tier "
                f"counties / geography drift — expected). Examples:"))
            for code, name in list(unmatched.items())[:8]:
                self.stdout.write(f"  {code}  {name}")

    def _fetch(self):
        self.stdout.write(f"Fetching ONS LE workbook from {ONS_URL} ...")
        try:
            with urllib.request.urlopen(
                    urllib.request.Request(ONS_URL, headers={"User-Agent": "Mozilla/5.0"}),
                    timeout=300) as resp:
                data = resp.read()
        except Exception as exc:
            raise CommandError(f"Could not fetch ONS LE workbook: {exc}")
        if len(data) < 10000:
            raise CommandError(f"ONS LE download too small ({len(data)} bytes) — URL moved?")
        return io.BytesIO(data)
