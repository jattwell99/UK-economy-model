"""
Ingest ONS regional GVA (balanced) — local authorities by ITL1/NUTS1 region.

Place at: core/management/commands/ingest_gva.py

Reads the ONS "Regional gross value added (balanced) by industry: local
authorities" workbooks (one file per ITL1 region). For each local authority it
takes the total-economy row (SIC07 = "Total", "All industries") from the
current-basic-prices sheet and writes one PlaceObservation per authority-year.

Usage:
    python manage.py ingest_gva --path /data/gva/           # ingest all region files in a dir
    python manage.py ingest_gva --path /data/gva/ --dry-run # parse + report, no writes
    python manage.py ingest_gva --path /data/gva/ --vintage 2019-12-19

Notes:
- The file glob deliberately matches only the "...localauthorities..." files, so
  the city-and-enterprise-regions workbook (different, non-LAD geography) is
  ignored even if it sits in the same folder.
- vintage is auto-detected from each file's Information sheet ("published on
  <date>"); override with --vintage if detection fails.
- Boundary churn is expected: this 2019 edition references the LAD set as it
  stood in 2019, so a few codes won't match a current Place set (authorities
  reorganised since). The command reports unmatched codes rather than dropping
  them silently — exactly the case Place versioning exists for.
"""

import glob
import os
import re
from datetime import date, datetime
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from core.models import (
    Indicator,
    PeriodType,
    Place,
    PlaceObservation,
    PlaceTier,
    Source,
)

FILE_GLOB = "regionalgrossvalueaddedbalancedbyindustrylocalauthorities*.xlsx"
SHEET = "Current Price"                 # current basic prices sheet (2019 edition)
# The 2025 edition renamed the sheets (Table 1-4) and drops "Current Price"; the
# current-price table is the one whose title says so. Code column also went
# "LAD code" -> "LA code". Both layouts are handled (see _resolve_price_sheet / CODE_COLS).
CODE_COLS = ("LAD code", "LA code")
INDICATOR_CODE = "gva-balanced-total"
SOURCE_NAME = "ONS Regional accounts (GVA / GDP / GDHI)"
SOURCE_PUBLISHER = "Office for National Statistics"


def _year_from_header(cell):
    """Header years can carry footnote suffixes, e.g. '20183' meaning 2018 note 3."""
    m = re.match(r"\s*(\d{4})", str(cell))
    if m:
        y = int(m.group(1))
        if 1990 <= y <= 2100:
            return y
    return None


def _detect_vintage(wb):
    """Read the 'published on <date>' line from the Information sheet."""
    if "Information" not in wb.sheetnames:
        return None
    for row in wb["Information"].iter_rows(max_row=20, values_only=True):
        for cell in row:
            if not cell:
                continue
            m = re.search(r"published on (\d{1,2} \w+ \d{4})", str(cell))
            if m:
                try:
                    return datetime.strptime(m.group(1), "%d %B %Y").date().isoformat()
                except ValueError:
                    return m.group(1)
    return None


def _resolve_price_sheet(wb, fname):
    """The current-basic-prices sheet, across edition layouts.

    2019 edition: a sheet literally named "Current Price". 2025 edition: the tables
    are "Table 1".."Table 4" and the current-price one is identified by its title
    text ("current prices"). Fall back to that title scan.
    """
    if SHEET in wb.sheetnames:
        return wb[SHEET]
    # Match on each sheet's OWN title cell (A1) — not any cell — so the "Contents"
    # sheet (which lists every table's title, including the current-price one) doesn't
    # win over the actual data table.
    for name in wb.sheetnames:
        ws = wb[name]
        title = next((row[0] for row in ws.iter_rows(min_row=1, max_row=1, values_only=True)), None)
        if title and "current price" in str(title).lower():
            return ws
    raise CommandError(f"{fname}: no current-price sheet (has {wb.sheetnames}).")


def parse_gva_totals(path):
    """Yield (gss_code, la_name, year, Decimal value) for total-economy rows."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = _resolve_price_sheet(wb, os.path.basename(path))

    header, hrow = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        if any(str(c).strip() in CODE_COLS for c in row if c is not None):
            header, hrow = list(row), i
            break
    if header is None:
        raise CommandError(f"{os.path.basename(path)}: no header row with a code column "
                           f"{CODE_COLS}.")

    col = {}
    for j, c in enumerate(header):
        s = str(c).strip() if c is not None else ""
        if s in CODE_COLS:
            col["code"] = j
        elif s == "LA name":
            col["name"] = j
        elif s == "SIC07":
            col["sic"] = j
        elif s == "SIC07 description":
            col["desc"] = j
    years = {j: y for j, c in enumerate(header) if (y := _year_from_header(c))}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hrow:
            continue
        sic = row[col["sic"]]
        if sic is None or str(sic).strip().lower() != "total":
            continue
        code = row[col["code"]]
        name = row[col["name"]]
        for j, yr in years.items():
            v = row[j] if j < len(row) else None
            if isinstance(v, (int, float)):
                yield code, name, yr, Decimal(str(v))

    vintage = _detect_vintage(wb)
    parse_gva_totals.last_vintage = vintage   # cheap side channel per file


class Command(BaseCommand):
    help = "Ingest ONS regional GVA (balanced) local-authority files into PlaceObservation."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=".",
                            help="Directory containing the region .xlsx files.")
        parser.add_argument("--vintage", default=None,
                            help="Override the auto-detected release vintage.")
        parser.add_argument("--dry-run", action="store_true",
                            help="Parse and report; write nothing.")

    def handle(self, *args, **opts):
        files = sorted(glob.glob(os.path.join(opts["path"], FILE_GLOB)))
        if not files:
            raise CommandError(f"No files matching {FILE_GLOB} under {opts['path']!r}.")

        try:
            indicator = Indicator.objects.get(code=INDICATOR_CODE)
        except Indicator.DoesNotExist:
            raise CommandError(f"Indicator '{INDICATOR_CODE}' not found — run seed_v1 --dimensions first.")
        source, _ = Source.objects.get_or_create(
            name=SOURCE_NAME, publisher=SOURCE_PUBLISHER,
        )

        # Cache current LAD places by code (latest boundary version).
        place_by_code = {}
        for p in Place.objects.filter(tier=PlaceTier.LAD).order_by("valid_from"):
            place_by_code[p.gss_code] = p   # later valid_from wins

        written = updated = skipped = 0
        unmatched = {}
        dry = opts["dry_run"]

        for path in files:
            rows = list(parse_gva_totals(path))
            vintage = opts["vintage"] or parse_gva_totals.last_vintage or "unknown"
            fname = os.path.basename(path)
            file_written = 0

            if dry:
                lads = {r[0] for r in rows}
                self.stdout.write(f"{fname}: {len(rows)} rows, {len(lads)} LADs, vintage={vintage}")
                continue

            with transaction.atomic():
                for code, name, year, value in rows:
                    place = place_by_code.get(code)
                    if place is None:
                        unmatched[code] = name
                        skipped += 1
                        continue
                    _, created = PlaceObservation.objects.update_or_create(
                        indicator=indicator,
                        place=place,
                        period_start=date(year, 1, 1),
                        period_end=date(year, 12, 31),
                        source=source,
                        vintage=vintage,
                        defaults={
                            "value": value,
                            "period_type": PeriodType.CALENDAR_YEAR,
                            "unit": "",          # inherit indicator.unit (£m)
                        },
                    )
                    written += int(created)
                    updated += int(not created)
                    file_written += 1
            self.stdout.write(f"{fname}: {file_written} observations (vintage {vintage}).")

        if dry:
            self.stdout.write(self.style.SUCCESS("Dry run complete."))
            return

        self.stdout.write(self.style.SUCCESS(
            f"Done. {written} created, {updated} updated, {skipped} skipped (unmatched)."
        ))
        if unmatched:
            self.stdout.write(self.style.WARNING(
                f"{len(unmatched)} GSS codes had no matching current LAD Place "
                f"(boundary churn — expected). Examples:"
            ))
            for code, name in list(unmatched.items())[:10]:
                self.stdout.write(f"  {code}  {name}")
