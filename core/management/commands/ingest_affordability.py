"""
Ingest ONS housing affordability — the house-price-to-residence-based-earnings ratio at
the LAD tier (England & Wales).

Why ingest ONS's ratio rather than derive it: ONS's figure is median-house-price over
median residence-based earnings — a like-for-like median-over-median ratio. Deriving from
our average-house-price (a mix-adjusted mean) and median-weekly-pay would mix a mean with
a median AND be GB-only (our pay indicator has no NI). ONS's ratio is the honest measure.

COVERAGE: England & Wales only. ONS does not produce this ratio for Scotland or Northern
Ireland — those nations publish their own (different construction), noted as a follow-up.
So this is one E&W-scoped indicator (PARTIAL_COVERAGE), comparable across E&W places — not
a per-nation split.

    python manage.py ingest_affordability                 # fetch current .xlsx from ONS
    python manage.py ingest_affordability --path aff.xlsx
    python manage.py ingest_affordability --dry-run

Reads sheet "5c" (Table 5c — "Ratio of median house price to median gross annual
residence-based earnings", local-authority level). Annual columns 2002..latest; the
trailing "5-Year Average" column is skipped (not a year). Suppressed cells ("[x]", e.g.
City of London) are skipped — they render as honest no-data. Codes match the current LAD
spine (including the 7 post-2019 unitaries); unmatched are reported, not dropped silently.
"""

import io
import urllib.request
from datetime import date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from core.models import (
    Indicator,
    ObservationStatus,
    PeriodType,
    Place,
    PlaceObservation,
    PlaceTier,
    Source,
)
# Barnsley/Sheffield recodes: this ONS product uses the new codes; map them to the spine.
from core.management.commands.refresh_lad_spine import RECODE_ALIASES

ONS_URL = (
    "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/housing/datasets/"
    "ratioofhousepricetoresidencebasedearningslowerquartileandmedian/current/"
    "aff2ratioofhousepricetoresidencebasedearnings.xlsx"
)
SHEET = "5c"                      # LA-level ratio of median house price to median earnings
CODE_HEADER = "Local authority code"
NAME_HEADER = "Local authority name"
INDICATOR_CODE = "house-price-to-earnings-ratio-residence"
DEFAULT_VINTAGE = "ons-aff-2025-03"
SOURCE_NAME = "ONS Housing affordability"
SOURCE_PUBLISHER = "Office for National Statistics"


def _year(cell):
    s = str(cell).strip() if cell is not None else ""
    if s[:4].isdigit() and 1990 <= int(s[:4]) <= 2100:
        return int(s[:4])
    return None   # e.g. "5-Year Average" -> skipped


def parse_affordability(path_or_fileobj):
    """Yield (gss_code, la_name, year, Decimal ratio) from sheet 5c."""
    wb = load_workbook(path_or_fileobj, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise CommandError(f"Affordability workbook has no sheet {SHEET!r} (has {wb.sheetnames}).")
    ws = wb[SHEET]

    header, hrow = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if CODE_HEADER in cells:
            header, hrow = cells, i
            break
    if header is None:
        raise CommandError(f"Affordability sheet {SHEET!r}: header row (with "
                           f"{CODE_HEADER!r}) not found.")
    code_j = header.index(CODE_HEADER)
    name_j = header.index(NAME_HEADER) if NAME_HEADER in header else code_j + 1
    years = {j: y for j, c in enumerate(header) if (y := _year(c))}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hrow:
            continue
        code = row[code_j] if code_j < len(row) else None
        if not code or not str(code).strip():
            continue
        code = str(code).strip()
        name = str(row[name_j]).strip() if name_j < len(row) and row[name_j] else ""
        for j, yr in years.items():
            v = row[j] if j < len(row) else None
            if not isinstance(v, (int, float)):
                continue   # suppressed "[x]" / blanks -> honest no-data
            try:
                yield code, name, yr, Decimal(str(v))
            except (InvalidOperation, ValueError):
                continue


class Command(BaseCommand):
    help = "Ingest ONS house-price-to-residence-earnings ratio (E&W, LAD) into PlaceObservation."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=None, help="Local .xlsx; else fetch from ONS.")
        parser.add_argument("--vintage", default=DEFAULT_VINTAGE)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        vintage = opts["vintage"]
        try:
            indicator = Indicator.objects.get(code=INDICATOR_CODE)
        except Indicator.DoesNotExist:
            raise CommandError(f"Indicator {INDICATOR_CODE!r} not seeded — run migrations / seed_v1.")
        source, _ = Source.objects.get_or_create(name=SOURCE_NAME, publisher=SOURCE_PUBLISHER)

        src = opts["path"] or self._fetch()
        rows = list(parse_affordability(src))
        if not rows:
            raise CommandError("Affordability sheet parsed to zero rows.")

        place_by_code = {
            p.gss_code: p for p in Place.objects.filter(tier=PlaceTier.LAD).order_by("valid_from")
        }
        objs, unmatched, kept = [], {}, 0
        for code, name, yr, value in rows:
            code = RECODE_ALIASES.get(code, code)   # Barnsley/Sheffield recodes -> spine code
            place = place_by_code.get(code)
            if place is None:
                unmatched[code] = name
                continue
            kept += 1
            objs.append(PlaceObservation(
                indicator=indicator, place=place,
                period_start=date(yr, 1, 1), period_end=date(yr, 12, 31),
                period_type=PeriodType.CALENDAR_YEAR,
                value=value, unit="", source=source, vintage=vintage,
                status=ObservationStatus.FINAL,
            ))

        if opts["dry_run"]:
            lads = {r[0] for r in rows}
            yrs = sorted({r[2] for r in rows})
            self.stdout.write(
                f"Affordability dry run: {kept} obs, {len(lads)} LADs, "
                f"years {yrs[0]}-{yrs[-1]}, {len(unmatched)} unmatched, vintage {vintage}.")
            return

        before = PlaceObservation.objects.filter(
            indicator=indicator, source=source, vintage=vintage).count()
        with transaction.atomic():
            for i in range(0, len(objs), 5000):
                PlaceObservation.objects.bulk_create(objs[i:i + 5000], ignore_conflicts=True)
        after = PlaceObservation.objects.filter(
            indicator=indicator, source=source, vintage=vintage).count()

        self.stdout.write(self.style.SUCCESS(
            f"Affordability (E&W): {after - before} created "
            f"({len({o.place_id for o in objs})} LADs), vintage {vintage}."))
        if unmatched:
            self.stdout.write(self.style.WARNING(
                f"{len(unmatched)} codes had no matching LAD Place "
                f"(geography drift). Examples:"))
            for code, name in list(unmatched.items())[:8]:
                self.stdout.write(f"  {code}  {name}")

    def _fetch(self):
        self.stdout.write(f"Fetching ONS affordability workbook from {ONS_URL} ...")
        try:
            with urllib.request.urlopen(
                    urllib.request.Request(ONS_URL, headers={"User-Agent": "Mozilla/5.0"}),
                    timeout=180) as resp:
                data = resp.read()
        except Exception as exc:
            raise CommandError(f"Could not fetch ONS affordability workbook: {exc}")
        if len(data) < 10000:
            raise CommandError(f"ONS affordability download too small ({len(data)} bytes) — URL moved?")
        return io.BytesIO(data)
