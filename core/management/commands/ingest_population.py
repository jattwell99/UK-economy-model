"""
Ingest ONS mid-year population estimates by local authority.

Place at: core/management/commands/ingest_population.py

Reads populationestimatesbylocalauthority.xlsx (single sheet, one row per LAD,
total persons by year) and writes one PlaceObservation per authority-year for
the `population` indicator. Population is a COUNT and IS additive (LAD populations
sum to regions), so it may be rolled up through the crosswalk.

Usage:
    python manage.py ingest_population --path /data/gva/populationestimatesbylocalauthority.xlsx
    python manage.py ingest_population --path <file> --vintage 2020-06-mye --dry-run

The workbook carries no release-date metadata sheet, so pass --vintage with the
ONS edition (e.g. the mid-year-estimates reference) for clean provenance;
defaults to "unknown" otherwise.
"""

import re
from datetime import date
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from core.models import (
    Indicator,
    IndicatorDomain,
    PeriodType,
    Place,
    PlaceObservation,
    PlaceTier,
    Source,
    SubjectScope,
    ValueType,
)

# The 2019 edition's sheet was "Data population"; the 2025 edition renamed it
# "Population data" and the code column "Region Code"/"LAD code" -> "LA code". Both handled.
SHEET_CANDIDATES = ("Data population", "Population data")
CODE_COLS = ("Region Code", "LAD code", "Code", "LA code")
INDICATOR_CODE = "population"
SOURCE_NAME = "ONS mid-year population estimates"
SOURCE_PUBLISHER = "Office for National Statistics"


def _year(cell):
    m = re.match(r"\s*(\d{4})", str(cell))
    if m and 1990 <= int(m.group(1)) <= 2100:
        return int(m.group(1))
    return None


def parse_population(path):
    """Yield (gss_code, name, year, Decimal count)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = next((wb[s] for s in SHEET_CANDIDATES if s in wb.sheetnames), wb.active)

    header, hrow = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        if any(str(c).strip() in CODE_COLS for c in row if c is not None):
            header, hrow = list(row), i
            break
    if header is None:
        raise CommandError(f"{path}: no header row found.")

    code_j = next(j for j, c in enumerate(header)
                  if str(c).strip() in CODE_COLS)
    name_j = code_j + 1
    years = {j: y for j, c in enumerate(header) if (y := _year(c))}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hrow:
            continue
        code = row[code_j]
        if not code:
            continue
        for j, yr in years.items():
            v = row[j] if j < len(row) else None
            if isinstance(v, (int, float)):
                yield code, row[name_j], yr, Decimal(str(v))


class Command(BaseCommand):
    help = "Ingest ONS LAD population estimates into PlaceObservation."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True)
        parser.add_argument("--vintage", default="unknown")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        indicator = self._get_or_create_indicator()
        source, _ = Source.objects.get_or_create(
            name=SOURCE_NAME, publisher=SOURCE_PUBLISHER,
        )
        rows = list(parse_population(opts["path"]))
        vintage = opts["vintage"]

        if opts["dry_run"]:
            lads = {r[0] for r in rows}
            self.stdout.write(f"{len(rows)} rows, {len(lads)} LADs, "
                              f"years {min(r[2] for r in rows)}-{max(r[2] for r in rows)}, "
                              f"vintage={vintage}")
            return

        place_by_code = {
            p.gss_code: p
            for p in Place.objects.filter(tier=PlaceTier.LAD).order_by("valid_from")
        }
        written = updated = skipped = 0
        unmatched = {}

        with transaction.atomic():
            for code, name, year, value in rows:
                place = place_by_code.get(code)
                if place is None:
                    unmatched[code] = name
                    skipped += 1
                    continue
                _, created = PlaceObservation.objects.update_or_create(
                    indicator=indicator, place=place,
                    period_start=date(year, 1, 1), period_end=date(year, 12, 31),
                    source=source, vintage=vintage,
                    defaults={"value": value, "period_type": PeriodType.CALENDAR_YEAR,
                              "unit": ""},
                )
                written += int(created)
                updated += int(not created)

        self.stdout.write(self.style.SUCCESS(
            f"Population: {written} created, {updated} updated, {skipped} skipped."
        ))
        if unmatched:
            self.stdout.write(self.style.WARNING(
                f"{len(unmatched)} codes unmatched (boundary churn). "
                f"e.g. {list(unmatched.items())[:5]}"
            ))

    @staticmethod
    def _get_or_create_indicator():
        try:
            return Indicator.objects.get(code=INDICATOR_CODE)
        except Indicator.DoesNotExist:
            domain, _ = IndicatorDomain.objects.get_or_create(
                code="demography", defaults={"name": "Demography"},
            )
            return Indicator.objects.create(
                code=INDICATOR_CODE, name="Population (mid-year estimate)",
                domain=domain, unit="count", value_type=ValueType.COUNT,
                is_additive=True, subject_scope=SubjectScope.PLACE,
            )
