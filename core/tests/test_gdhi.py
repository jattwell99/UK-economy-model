"""
Phase 3 breadth — Source 2: GDHI (economy).

Covers the two-table ingester (total £m + per-head £, per-head taken from ONS not
derived), CALENDAR_YEAR periods with a provisional latest year, and the additive /
non-additive roll-up behaviour.
"""

import tempfile
from datetime import date
from decimal import Decimal

from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook

from core.aggregation import NonAdditiveRollupError, rollup_place_value
from core.models import (
    CrosswalkBasis,
    Indicator,
    ObservationStatus,
    PeriodType,
    PlaceObservation,
    PlaceTier,
    ValueType,
)

from .factories import (
    make_crosswalk,
    make_domain,
    make_indicator,
    make_observation,
    make_place,
    make_source,
)


def _write_gdhi_workbook(path):
    wb = Workbook()
    wb.remove(wb.active)

    info = wb.create_sheet("Information")
    info.append(["Regional gross disposable household income (GDHI): local authorities"])
    info.append(["These tables are part of the regional GDHI release published on 4 September 2024."])

    t1 = wb.create_sheet("Table 1")
    t1.append(["Table 1: GDHI local authority: total GDHI at current basic prices, pounds million"])
    t1.append(["Region", "LAD code", "Region name", "2021", "2022"])
    t1.append(["North East", "E06000001", "Hartlepool", 1700, 1750])
    t1.append(["North West", "E08000003", "Manchester", 9269, 9984])

    t3 = wb.create_sheet("Table 3")
    t3.append(["Table 3: GDHI local authority: GDHI per head of population at current basic prices, pounds"])
    t3.append(["Region", "LAD code", "Region name", "2021", "2022"])
    t3.append(["North East", "E06000001", "Hartlepool", 17500, 17992])
    t3.append(["North West", "E08000003", "Manchester", 16800, 17615])

    # A decoy per-head *indices* sheet that must NOT be picked as per-head £.
    t4 = wb.create_sheet("Table 4")
    t4.append(["Table 4: GDHI local authority: GDHI per head indices (UK=100)"])
    t4.append(["Region", "LAD code", "Region name", "2021", "2022"])
    t4.append(["North West", "E08000003", "Manchester", 88, 89])

    wb.save(path)


class GdhiIngestTests(TestCase):
    def setUp(self):
        make_place("E06000001", "Hartlepool", tier=PlaceTier.LAD)
        make_place("E08000003", "Manchester", tier=PlaceTier.LAD)

    def _ingest(self):
        with tempfile.NamedTemporaryFile("wb", suffix=".xlsx", delete=False) as fh:
            path = fh.name
        _write_gdhi_workbook(path)
        call_command("ingest_gdhi", path=path)

    def test_creates_total_and_perhead(self):
        self._ingest()
        total = Indicator.objects.get(code="gdhi-total")
        perhead = Indicator.objects.get(code="gdhi-per-head")
        self.assertTrue(total.is_additive)
        self.assertEqual(total.value_type, ValueType.CURRENCY)
        self.assertFalse(perhead.is_additive)

        # 2 LADs x 2 years x 2 indicators = 8 observations (indices sheet ignored)
        self.assertEqual(PlaceObservation.objects.filter(
            indicator__in=[total, perhead]).count(), 8)

        o = PlaceObservation.objects.get(
            indicator=perhead, place__gss_code="E08000003", period_start=date(2022, 1, 1))
        self.assertEqual(o.value, Decimal("17615"))
        self.assertEqual(o.period_type, PeriodType.CALENDAR_YEAR)
        self.assertEqual(o.period_end, date(2022, 12, 31))
        self.assertEqual(o.vintage, "2024-09-04")

    def test_latest_year_is_provisional(self):
        self._ingest()
        total = Indicator.objects.get(code="gdhi-total")
        y2022 = PlaceObservation.objects.get(
            indicator=total, place__gss_code="E08000003", period_start=date(2022, 1, 1))
        y2021 = PlaceObservation.objects.get(
            indicator=total, place__gss_code="E08000003", period_start=date(2021, 1, 1))
        self.assertEqual(y2022.status, ObservationStatus.PROVISIONAL)
        self.assertEqual(y2021.status, "")

    def test_idempotent(self):
        self._ingest()
        self._ingest()
        self.assertEqual(PlaceObservation.objects.filter(
            indicator__code__in=["gdhi-total", "gdhi-per-head"]).count(), 8)


class GdhiRollupTests(TestCase):
    def setUp(self):
        self.economy = make_domain("economy", "Economy")
        self.src = make_source()
        self.lad_a = make_place("E06000001", "A", tier=PlaceTier.LAD)
        self.lad_b = make_place("E08000003", "B", tier=PlaceTier.LAD)
        self.wpc = make_place("E14000001", "C", tier=PlaceTier.WPC)
        make_crosswalk(self.lad_a, self.wpc, Decimal("1.000000"), CrosswalkBasis.POPULATION)
        make_crosswalk(self.lad_b, self.wpc, Decimal("0.500000"), CrosswalkBasis.POPULATION)

    def test_gdhi_total_rolls_up(self):
        tot = make_indicator(self.economy, code="gdhi-total",
                             is_additive=True, value_type=ValueType.CURRENCY, unit="£m")
        make_observation(tot, self.lad_a, self.src, value=Decimal("1000"),
                         period_start=date(2022, 1, 1), period_end=date(2022, 12, 31))
        make_observation(tot, self.lad_b, self.src, value=Decimal("2000"),
                         period_start=date(2022, 1, 1), period_end=date(2022, 12, 31))
        total = rollup_place_value(
            tot, self.wpc, period_start=date(2022, 1, 1), period_end=date(2022, 12, 31),
            basis=CrosswalkBasis.POPULATION)
        self.assertEqual(total, Decimal("2000"))  # 1000*1.0 + 2000*0.5

    def test_gdhi_per_head_refuses(self):
        ph = make_indicator(self.economy, code="gdhi-per-head",
                            is_additive=False, value_type=ValueType.RATIO, unit="£")
        make_observation(ph, self.lad_a, self.src, value=Decimal("20000"),
                         period_start=date(2022, 1, 1), period_end=date(2022, 12, 31))
        with self.assertRaises(NonAdditiveRollupError):
            rollup_place_value(
                ph, self.wpc, period_start=date(2022, 1, 1), period_end=date(2022, 12, 31),
                basis=CrosswalkBasis.POPULATION)
