"""
Staleness refresh — the GVA / population parsers must handle BOTH ONS edition
layouts (docs: economy staleness refresh to the 2025 editions).

The 2025 edition moved the GVA current-price data to a "Table N" sheet (no more
"Current Price"), renamed the code column "LAD code" -> "LA code", and the
population sheet "Data population" -> "Population data". These fixtures reproduce
both layouts so a future edition-shift is caught here, not in production.
"""

import tempfile

from django.test import SimpleTestCase

from openpyxl import Workbook

from core.management.commands.ingest_gva import parse_gva_totals
from core.management.commands.ingest_population import parse_population


def _save(wb):
    fh = tempfile.NamedTemporaryFile("wb", suffix=".xlsx", delete=False)
    wb.save(fh.name)
    return fh.name


class GvaEditionDriftTests(SimpleTestCase):
    def test_2025_layout_table3_la_code(self):
        wb = Workbook()
        # A "Contents" sheet that LISTS the current-price table title — the resolver
        # must not mistake it for the data table (it comes first).
        contents = wb.active
        contents.title = "Contents"
        contents.append(["Contents"])
        contents.append(["Table 3", "Table 3: ... current prices, pounds million"])
        # Decoy chained-volume table before the real one.
        t2 = wb.create_sheet("Table 2")
        t2.append(["Table 2: ... chained volume measures, pounds million"])
        t3 = wb.create_sheet("Table 3")
        t3.append(["Table 3: ... current prices, pounds million"])
        t3.append(["ITL1 region", "LA code", "LA name", "SIC07", "SIC07 description", 2022, 2023])
        t3.append(["North East", "E06000001", "Hartlepool", "Total", "All industries", 1256, 1353])
        t3.append(["North East", "E06000001", "Hartlepool", "A-E", "Production", 300, 320])
        rows = list(parse_gva_totals(_save(wb)))
        # Only the Total row, and it resolved Table 3 (not Contents/Table 2).
        self.assertIn(("E06000001", "Hartlepool", 2023, 1353), [(c, n, y, int(v)) for c, n, y, v in rows])
        self.assertEqual(len(rows), 2)  # 2022 + 2023 of the Total row only

    def test_2019_layout_still_parses(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Current Price"
        ws.append(["LAD code", "LA name", "SIC07", "SIC07 description", 2017, 2018])
        ws.append(["E06000001", "Hartlepool", "Total", "All industries", 1400, 1457])
        ws.append(["E06000001", "Hartlepool", "C", "Manufacturing", 200, 201])
        rows = [(c, n, y, int(v)) for c, n, y, v in parse_gva_totals(_save(wb))]
        self.assertIn(("E06000001", "Hartlepool", 2018, 1457), rows)
        self.assertEqual(len(rows), 2)


class PopulationEditionDriftTests(SimpleTestCase):
    def test_2025_layout_population_data_la_code(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Population data"
        ws.append(["ITL1 Region", "LA code", "LA name", 2022, 2023])
        ws.append(["North East", "E06000001", "Hartlepool", 90000, 91000])
        rows = [(c, n, y, int(v)) for c, n, y, v in parse_population(_save(wb))]
        self.assertIn(("E06000001", "Hartlepool", 2023, 91000), rows)

    def test_2019_layout_still_parses(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data population"
        ws.append(["LAD code", "LA name", 2018, 2019])
        ws.append(["E06000001", "Hartlepool", 88000, 89000])
        rows = [(c, n, y, int(v)) for c, n, y, v in parse_population(_save(wb))]
        self.assertIn(("E06000001", "Hartlepool", 2019, 89000), rows)
